import datetime as dt
import sqlite3
from typing import List, Iterable, Dict, Optional

from openpyxl import Workbook, load_workbook

from anpy import AbstractDataHandler
from anpy import Record
from anpy_lib import column_creation as cc, data_handling, data_analysis

TEMP_SHEET_NAME = 'ANPY_TEMP_SHEET_DO_NOT_TOUCH'


def enter_week_data(first: dt.datetime, handler: AbstractDataHandler, ws):
    weekly_record_list = data_analysis.get_records_on_week(handler, first)

    dicts = data_analysis.get_total_subject_breakdown(weekly_record_list)
    make_cols(first, weekly_record_list, dicts)
    cc.Column.make_all(ws)


def make_cols(first: dt.datetime, weekly_record_list, dicts):
    starts, ends = get_time_started_ended(weekly_record_list)
    cc.DateColumn(first)
    cc.TimeStartedColumn(starts)
    cc.TimeEndedColumn(ends)
    cc.TimeTotalColumn()
    cc.TimeWorkingColumn()
    cc.EfficiencyColumn()
    a = list(get_data_column_data(dicts).items())
    for sub, times in get_data_column_data(dicts).items():
        cc.CategoryTimeColumn(sub, [t / 60 if t else t for t in times])


def get_time_started_ended(weekly_record_list: List[Iterable[Record]]):
    starts = []
    ends = []

    for records in weekly_record_list:
        records = list(records)
        if records:
            start, end = records[0].start, records[-1].end
            starts.append(start.time() if start else None)
            ends.append(end.time() if start else None)
        else:
            starts.append(None)
            ends.append(None)
    return starts, ends


def get_data_column_data(dicts: List[Dict[str, float]]) \
        -> Dict[str, List[Optional[float]]]:
    # list per day
    default_col = [None] * 7

    return_dict: Dict[str, List[Optional[float]]] = dict()

    for idx, d in enumerate(dicts):
        for subject, time in d.items():
            if subject not in return_dict:
                return_dict[subject] = list(default_col)
            return_dict[subject][idx] = time

    return return_dict


def get_most_recent_monday(datetime: dt.datetime = None):
    if not datetime:
        datetime = dt.datetime.today()
    datetime = datetime - dt.timedelta(hours=6)
    datetime = datetime - dt.timedelta(days=datetime.isoweekday())
    return dt.datetime.combine(datetime.date(), dt.time(6, 0))


def load_excel_workbook(path):
    try:
        wb = load_workbook(path)
    except FileNotFoundError:
        wb = Workbook()
        wb.active.title = TEMP_SHEET_NAME
    return wb


def get_relevant_worksheet(workbook: Workbook, datetime=None):
    reference_date = str(get_most_recent_monday(datetime).date())
    if reference_date not in workbook.sheetnames:
        workbook.create_sheet(title=reference_date)
    else:
        workbook[reference_date].title = TEMP_SHEET_NAME
        workbook.create_sheet(title=reference_date)

    if TEMP_SHEET_NAME in workbook.sheetnames:
        del workbook[TEMP_SHEET_NAME]
    return workbook[reference_date], get_most_recent_monday(datetime)


if __name__ == '__main__':
    a = Workbook()
    ws = a.active
    h = data_handling.SQLDataHandler(sqlite3.Connection('test.db'))
    h.new_category('a')
    h.new_category('b')
    h.new_category('c')
    cats = h.get_categories()
    h.start(cats['a'], dt.datetime(2015, 3, 3, 8, 30))
    h.complete(dt.datetime(2015, 3, 3, 10, 0))
    h.start(cats['a'], dt.datetime(2015, 3, 3, 13, 0))
    h.complete(dt.datetime(2015, 3, 3, 15, 0))
    h.start(cats['b'], dt.datetime(2015, 3, 4, 14, 0))
    h.complete(dt.datetime(2015, 3, 4, 15, 0))

    enter_week_data(dt.datetime(2015, 2, 28, 6, 0), h, ws)
    a.save('done.xlsx')
